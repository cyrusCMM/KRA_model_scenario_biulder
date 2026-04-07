# -*- coding: utf-8 -*-
import shutil
from pathlib import Path
from tempfile import NamedTemporaryFile

import altair as alt
import pandas as pd
import streamlit as st
from openpyxl import load_workbook

from app_runner_v3 import run_app
from export_engine_v1 import export_full_package


FISCAL_ORDER = ["Jul", "Aug", "Sep", "Oct", "Nov", "Dec", "Jan", "Feb", "Mar", "Apr", "May", "Jun"]

st.set_page_config(page_title="KRA Revenue Forecasting Model", layout="wide")
st.title("KRA Revenue Forecasting Model")


# ============================================================
# BASIC HELPERS
# ============================================================

def _clean(x):
    if pd.isna(x):
        return ""
    return str(x).strip()


def next_fy_label(year_label: str) -> str:
    start, end = year_label.split("/")
    return f"{int(start) + 1}/{int(end) + 1}"


@st.cache_data(show_spinner=False)
def load_workbook_metadata(workbook_path: str):
    control = pd.read_excel(workbook_path, sheet_name="Control", header=None)
    scenarios = pd.read_excel(workbook_path, sheet_name="CGE_Scenarios")

    selected_year = None
    selected_scenario = None

    for i in range(len(control)):
        key = _clean(control.iloc[i, 0])
        val = control.iloc[i, 1] if control.shape[1] > 1 else None
        if key == "Selected Year":
            selected_year = _clean(val)
        elif key == "Selected Scenario":
            selected_scenario = _clean(val)

    scenario_names = []
    if "Scenario" in scenarios.columns:
        scenario_names = scenarios["Scenario"].dropna().astype(str).str.strip().tolist()

    return {
        "selected_year": selected_year,
        "selected_scenario": selected_scenario,
        "scenarios_df": scenarios,
        "scenario_names": scenario_names,
    }


def _set_control_value(ws, label: str, value):
    for row in ws.iter_rows():
        for cell in row:
            if _clean(cell.value) == label:
                ws.cell(row=cell.row, column=cell.column + 1).value = value
                return

    next_row = ws.max_row + 1
    ws.cell(row=next_row, column=1).value = label
    ws.cell(row=next_row, column=2).value = value


def _override_scenario_row(ws, scenario_name: str, scenario_row_df: pd.DataFrame):
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    header_map = {_clean(h): idx + 1 for idx, h in enumerate(headers)}

    if "Scenario" not in headers:
        raise ValueError("CGE_Scenarios sheet missing 'Scenario' header.")

    scenario_col = headers.index("Scenario") + 1
    target_row = None

    for r in range(2, ws.max_row + 1):
        if _clean(ws.cell(row=r, column=scenario_col).value) == _clean(scenario_name):
            target_row = r
            break

    if target_row is None:
        raise ValueError(f"Scenario '{scenario_name}' not found in CGE_Scenarios.")

    row_dict = scenario_row_df.iloc[0].to_dict()
    for k, v in row_dict.items():
        k_clean = _clean(k)
        if k_clean in header_map:
            ws.cell(row=target_row, column=header_map[k_clean]).value = v


def prepare_temp_workbook(
    source_path: str | Path,
    selected_year: str,
    selected_scenario: str,
    mode: str,
    custom_scenario_row: pd.DataFrame | None = None,
) -> Path:
    source_path = Path(source_path).resolve()

    tmp = NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.close()
    temp_path = Path(tmp.name)

    shutil.copy2(source_path, temp_path)

    wb = load_workbook(temp_path)
    ws_control = wb["Control"]

    _set_control_value(ws_control, "Selected Year", selected_year)
    _set_control_value(ws_control, "Selected Scenario", selected_scenario)
    _set_control_value(ws_control, "Mode", mode)
    _set_control_value(ws_control, "Current Fiscal Year", selected_year)

    if custom_scenario_row is not None:
        ws_scen = wb["CGE_Scenarios"]
        _override_scenario_row(ws_scen, selected_scenario, custom_scenario_row)

    wb.save(temp_path)
    wb.close()
    return temp_path


def run_package(
    workbook_path: str | Path,
    selected_year: str,
    selected_scenario: str,
    custom_scenario_row: pd.DataFrame | None = None,
    mode: str = "scenario",
):
    temp_path = prepare_temp_workbook(
        source_path=workbook_path,
        selected_year=selected_year,
        selected_scenario=selected_scenario,
        mode=mode,
        custom_scenario_row=custom_scenario_row,
    )
    try:
        result = run_app(
            workbook_path=temp_path,
            mode=mode,
            export=False,
        )
        return result
    finally:
        try:
            temp_path.unlink(missing_ok=True)
        except Exception:
            pass


# ============================================================
# TWO-FY PRESENTATION HELPERS
# ============================================================

def build_two_fy_monthly_bridge(current_monthly: pd.DataFrame, next_monthly: pd.DataFrame | None = None) -> pd.DataFrame:
    def _prep(df, fy_rank):
        x = df.copy()
        x["Month Index"] = pd.to_numeric(x["Month Index"], errors="coerce")

        for c in ["Baseline Monthly Forecast", "Scenario Monthly Forecast", "Monthly Impact", "Actual Monthly"]:
            if c not in x.columns:
                x[c] = 0.0
            x[c] = pd.to_numeric(x[c], errors="coerce").fillna(0.0)

        x = x.groupby(
            ["Fiscal Year", "Month Index", "Month Name"],
            as_index=False
        )[[
            "Baseline Monthly Forecast",
            "Scenario Monthly Forecast",
            "Monthly Impact",
            "Actual Monthly"
        ]].sum()

        x["FY Rank"] = fy_rank
        x["Month Sequence"] = (fy_rank - 1) * 12 + x["Month Index"]
        x["Display Label"] = x["Fiscal Year"].astype(str) + "-" + x["Month Name"].astype(str)

        x["Actual Share Flag"] = x["Actual Monthly"].abs() > 0

        return x

    out = [_prep(current_monthly, 1)]
    if next_monthly is not None and not next_monthly.empty:
        out.append(_prep(next_monthly, 2))

    df = pd.concat(out, ignore_index=True)
    return df.sort_values("Month Sequence").reset_index(drop=True)


def build_two_fy_annual_comparison(current_result: dict, next_result: dict | None = None) -> pd.DataFrame:
    rows = []

    cur = current_result["comparisons"]["total_comparison"].iloc[0]
    rows.append({
        "Fiscal Year": current_result["metadata"]["selected_year"],
        "Baseline Final": float(cur["Baseline Final"]),
        "Scenario Final": float(cur["Scenario Final"]),
        "Scenario Impact": float(cur["Scenario Impact"]),
        "Scenario Impact %": float(cur["Scenario Impact %"]),
    })

    if next_result is not None:
        nxt = next_result["comparisons"]["total_comparison"].iloc[0]
        rows.append({
            "Fiscal Year": next_result["metadata"]["selected_year"],
            "Baseline Final": float(nxt["Baseline Final"]),
            "Scenario Final": float(nxt["Scenario Final"]),
            "Scenario Impact": float(nxt["Scenario Impact"]),
            "Scenario Impact %": float(nxt["Scenario Impact %"]),
        })

    return pd.DataFrame(rows)


def prepare_tax_head_impact_chart_df(tax_head_comparison: pd.DataFrame, top_n: int = 20) -> pd.DataFrame:
    df = tax_head_comparison.copy()
    df["Scenario Impact"] = pd.to_numeric(df["Scenario Impact"], errors="coerce").fillna(0.0)
    df = df.sort_values("Scenario Impact").tail(top_n).copy()
    return df


# ============================================================
# CHART BUILDERS
# ============================================================

def build_annual_grouped_chart(df: pd.DataFrame) -> alt.Chart:
    base = alt.Chart(df)

    baseline_bars = base.mark_bar(size=36).encode(
        x=alt.X("Fiscal Year:N", title="Fiscal Year"),
        y=alt.Y("Baseline Final:Q", title="Revenue"),
        color=alt.value("#4C78A8"),
        tooltip=[
            "Fiscal Year",
            alt.Tooltip("Baseline Final:Q", format=",.2f"),
        ],
    )

    scenario_bars = base.mark_bar(size=18).encode(
        x=alt.X("Fiscal Year:N"),
        y=alt.Y("Scenario Final:Q"),
        color=alt.value("#F58518"),
        tooltip=[
            "Fiscal Year",
            alt.Tooltip("Scenario Final:Q", format=",.2f"),
            alt.Tooltip("Scenario Impact:Q", format=",.2f"),
            alt.Tooltip("Scenario Impact %:Q", format=".2%"),
        ],
    )

    impact_text = base.mark_text(
        dy=-10,
        fontSize=12,
        color="black"
    ).encode(
        x=alt.X("Fiscal Year:N"),
        y=alt.Y("Scenario Final:Q"),
        text=alt.Text("Scenario Impact:Q", format=",.0f")
    )

    return (baseline_bars + scenario_bars + impact_text).properties(height=340)


def build_two_fy_line_chart(df: pd.DataFrame) -> alt.Chart:
    long_df = df.melt(
        id_vars=["Display Label", "MonthSequence"] if "MonthSequence" in df.columns else ["Display Label", "Month Sequence"],
        value_vars=["Baseline Monthly Forecast", "Scenario Monthly Forecast"],
        var_name="Series",
        value_name="Revenue",
    )

    seq_col = "MonthSequence" if "MonthSequence" in long_df.columns else "Month Sequence"
    sort_order = df.sort_values(seq_col)["Display Label"].tolist()

    base = alt.Chart(long_df).encode(
        x=alt.X("Display Label:N", sort=sort_order, title="Month"),
        y=alt.Y("Revenue:Q", title="Revenue"),
        color=alt.Color("Series:N", title="Series"),
        tooltip=[
            "Display Label",
            "Series",
            alt.Tooltip("Revenue:Q", format=",.2f"),
        ],
    )

    lines = base.mark_line(point=True)

    impact_line = (
        alt.Chart(df)
        .mark_line(strokeDash=[4, 2], color="black")
        .encode(
            x=alt.X("Display Label:N", sort=sort_order),
            y=alt.Y("Monthly Impact:Q", title="Revenue / Impact"),
            tooltip=[
                "Display Label",
                alt.Tooltip("Monthly Impact:Q", format=",.2f"),
            ],
        )
    )

    return (lines + impact_line).properties(height=380)


def build_two_fy_impact_chart(df: pd.DataFrame) -> alt.Chart:
    seq_col = "MonthSequence" if "MonthSequence" in df.columns else "Month Sequence"
    sort_order = df.sort_values(seq_col)["Display Label"].tolist()

    return (
        alt.Chart(df)
        .mark_bar()
        .encode(
            x=alt.X("Display Label:N", sort=sort_order, title="Month"),
            y=alt.Y("Monthly Impact:Q", title="Scenario - Baseline"),
            tooltip=[
                "Display Label",
                alt.Tooltip("Monthly Impact:Q", format=",.2f"),
            ],
        )
        .properties(height=260)
    )


def build_tax_head_impact_chart(df: pd.DataFrame) -> alt.Chart:
    df = df.sort_values("Scenario Impact").copy()
    sort_order = df["Internal Tax Head"].tolist()

    return (
        alt.Chart(df)
        .mark_bar()
        .encode(
            x=alt.X("Scenario Impact:Q", title="Impact"),
            y=alt.Y("Internal Tax Head:N", sort=sort_order, title="Tax Head"),
            tooltip=[
                "Internal Tax Head",
                alt.Tooltip("Scenario Impact:Q", format=",.2f"),
            ],
        )
        .properties(height=500)
    )


# ============================================================
# DISPLAY HELPERS
# ============================================================

def render_metrics(total_comparison: pd.DataFrame):
    row = total_comparison.iloc[0]
    c1, c2, c3 = st.columns(3)
    c1.metric("Baseline Final", f"{float(row['Baseline Final']):,.0f}")
    c2.metric("Scenario Final", f"{float(row['Scenario Final']):,.0f}")
    c3.metric("Scenario Impact", f"{float(row['Scenario Impact']):,.0f}", f"{float(row['Scenario Impact %']) * 100:,.2f}%")


# ============================================================
# SIDEBAR
# ============================================================

st.sidebar.header("Run Controls")

default_workbook = str(Path("kra_forecast_input_template_final.xlsx").resolve())
workbook_path = st.sidebar.text_input("Workbook path", value=default_workbook)

meta = None
if Path(workbook_path).exists():
    try:
        meta = load_workbook_metadata(workbook_path)
    except Exception as e:
        st.sidebar.error(f"Workbook read error: {e}")

run_type = st.sidebar.radio("Scenario mode", ["Official", "Custom"], horizontal=True)

selected_year = st.sidebar.selectbox(
    "Selected Year",
    options=["2025/26", "2026/27", "2027/28"],
    index=0 if meta is None or meta["selected_year"] not in ["2025/26", "2026/27", "2027/28"]
    else ["2025/26", "2026/27", "2027/28"].index(meta["selected_year"]),
)

scenario_options = meta["scenario_names"] if meta is not None and meta["scenario_names"] else ["Baseline"]
selected_scenario = st.sidebar.selectbox(
    "Selected Scenario",
    options=scenario_options,
    index=0 if meta is None or meta["selected_scenario"] not in scenario_options
    else scenario_options.index(meta["selected_scenario"]),
)

custom_row = None
if run_type == "Custom" and meta is not None:
    st.sidebar.caption("Custom mode edits a temporary copy of the selected scenario row. The workbook is not overwritten.")
    base_row = meta["scenarios_df"].loc[
        meta["scenarios_df"]["Scenario"].astype(str).str.strip() == selected_scenario
    ].copy()

    if base_row.empty:
        base_row = meta["scenarios_df"].head(1).copy()

    editable_cols = [c for c in base_row.columns if c not in ["Scenario", "Description", "Notes"]]
    editor_df = base_row[["Scenario", "Description"] + editable_cols].reset_index(drop=True)
    custom_row = st.sidebar.data_editor(editor_df, use_container_width=True, num_rows="fixed")

run_button = st.sidebar.button("Run Model", type="primary")


# ============================================================
# MAIN RUN
# ============================================================

if run_button:
    try:
        if not Path(workbook_path).exists():
            st.error("Workbook path not found.")
            st.stop()

        scenario_row_df = custom_row if run_type == "Custom" else None

        current_result = run_package(
            workbook_path=workbook_path,
            selected_year=selected_year,
            selected_scenario=selected_scenario,
            custom_scenario_row=scenario_row_df,
            mode="scenario",
        )

        next_result = None
        try:
            next_year = next_fy_label(selected_year)
            if next_year in ["2026/27", "2027/28"]:
                next_result = run_package(
                    workbook_path=workbook_path,
                    selected_year=next_year,
                    selected_scenario=selected_scenario,
                    custom_scenario_row=scenario_row_df,
                    mode="scenario",
                )
        except Exception as next_err:
            st.warning(f"Next-FY scenario run not available: {next_err}")

        st.success("Model run complete.")

        render_metrics(current_result["comparisons"]["total_comparison"])

        annual_2fy = build_two_fy_annual_comparison(current_result, next_result)

        st.subheader("Baseline vs Scenario — Annual Path")
        st.altair_chart(build_annual_grouped_chart(annual_2fy), use_container_width=True)

        current_monthly = current_result["dashboard_pack"]["monthly_total_comparison"].copy()
        next_monthly = None if next_result is None else next_result["dashboard_pack"]["monthly_total_comparison"].copy()
        monthly_2fy = build_two_fy_monthly_bridge(current_monthly, next_monthly)

        st.subheader("Baseline vs Scenario — Time-Varying Monthly Path")
        st.altair_chart(build_two_fy_line_chart(monthly_2fy), use_container_width=True)

        st.subheader("Scenario Impact by Month")
        st.altair_chart(build_two_fy_impact_chart(monthly_2fy), use_container_width=True)

        tax_head_chart_df = prepare_tax_head_impact_chart_df(
            current_result["comparisons"]["tax_head_comparison"],
            top_n=20,
        )
        st.subheader("Tax Head Scenario Impact")
        st.altair_chart(build_tax_head_impact_chart(tax_head_chart_df), use_container_width=True)

        c1, c2 = st.columns(2)
        with c1:
            st.subheader("Top Gainers")
            st.dataframe(current_result["dashboard_pack"]["top_gainers"], use_container_width=True)
        with c2:
            st.subheader("Top Losers")
            st.dataframe(current_result["dashboard_pack"]["top_losers"], use_container_width=True)

        st.subheader("Total Comparison")
        st.dataframe(current_result["comparisons"]["total_comparison"], use_container_width=True)

        st.subheader("Tax Head Comparison")
        st.dataframe(current_result["comparisons"]["tax_head_comparison"], use_container_width=True)

        st.subheader("Monthly Comparison — Current FY")
        st.dataframe(current_result["dashboard_pack"]["monthly_total_comparison"], use_container_width=True)

        if next_result is not None:
            st.subheader("Monthly Comparison — Next FY")
            st.dataframe(next_result["dashboard_pack"]["monthly_total_comparison"], use_container_width=True)

        augmented_dashboard_pack = dict(current_result["dashboard_pack"])
        augmented_dashboard_pack["two_fy_annual_comparison"] = annual_2fy
        augmented_dashboard_pack["two_fy_monthly_comparison"] = monthly_2fy

        st.header("Download Output")
        if st.button("Generate Excel Export"):
            output_file = Path("streamlit_export.xlsx").resolve()

            export_full_package(
                filepath=output_file,
                baseline_outputs=current_result["baseline"],
                scenario_outputs=current_result["scenario"],
                comparisons=current_result["comparisons"],
                dashboard_pack=augmented_dashboard_pack,
                decomposition_pack=current_result["decomposition_pack"],
                metadata=current_result["metadata"],
            )

            with open(output_file, "rb") as f:
                st.download_button(
                    label="Download Excel",
                    data=f,
                    file_name="kra_forecast_output.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

    except Exception as e:
        st.error(f"Error: {e}")