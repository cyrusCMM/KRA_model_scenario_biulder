# -*- coding: utf-8 -*-
"""
Created on Wed Apr  1 20:18:24 2026

@author: hp
"""
# -*- coding: utf-8 -*-
import traceback
from pathlib import Path
from tempfile import NamedTemporaryFile
import shutil

import pandas as pd
from openpyxl import load_workbook

from app_runner_v3 import run_app


WORKBOOK = r"C:\Users\hp\Documents\Q\KRA model\kra_forecast_model\inputs\kra_forecast_input_template_final.xlsx"


def _clean(x):
    if pd.isna(x):
        return ""
    return str(x).strip()


def _set_control_value(ws, label: str, value):
    for row in ws.iter_rows():
        for cell in row:
            if _clean(cell.value) == label:
                ws.cell(row=cell.row, column=cell.column + 1).value = value
                return
    next_row = ws.max_row + 1
    ws.cell(row=next_row, column=1).value = label
    ws.cell(row=next_row, column=2).value = value


def prepare_temp_workbook(source_path: str, selected_year: str, selected_scenario: str, mode: str = "scenario") -> Path:
    tmp = NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.close()
    temp_path = Path(tmp.name)

    shutil.copy2(source_path, temp_path)

    wb = load_workbook(temp_path)
    ws_control = wb["Control"]

    _set_control_value(ws_control, "Selected Year", selected_year)
    _set_control_value(ws_control, "Selected Scenario", selected_scenario)
    _set_control_value(ws_control, "Mode", mode)
    _set_control_value(ws_control, "Current Fiscal Year", selected_year)

    wb.save(temp_path)
    wb.close()
    return temp_path


def get_scenarios(workbook_path: str):
    df = pd.read_excel(workbook_path, sheet_name="CGE_Scenarios")
    if "Scenario" not in df.columns:
        raise ValueError("CGE_Scenarios sheet missing 'Scenario' column.")
    return [x for x in df["Scenario"].dropna().astype(str).str.strip().tolist() if x]


def run_one(workbook_path: str, selected_year: str, scenario_name: str):
    temp_path = prepare_temp_workbook(
        source_path=workbook_path,
        selected_year=selected_year,
        selected_scenario=scenario_name,
        mode="scenario",
    )
    try:
        result = run_app(
            workbook_path=temp_path,
            mode="scenario",
            export=False,
        )
        return result
    finally:
        try:
            temp_path.unlink(missing_ok=True)
        except Exception:
            pass


def summarize_result(result: dict, scenario_name: str):
    total = result["comparisons"]["total_comparison"].iloc[0]
    monthly = result["dashboard_pack"]["monthly_total_comparison"].copy()
    tax = result["comparisons"]["tax_head_comparison"].copy()

    monthly["Monthly Impact"] = pd.to_numeric(monthly["Monthly Impact"], errors="coerce").fillna(0.0)
    tax["Scenario Impact"] = pd.to_numeric(tax["Scenario Impact"], errors="coerce").fillna(0.0)

    top_gain = tax.sort_values("Scenario Impact", ascending=False).head(3)
    top_loss = tax.sort_values("Scenario Impact", ascending=True).head(3)

    flags = []

    baseline_final = float(total["Baseline Final"])
    scenario_final = float(total["Scenario Final"])
    scenario_impact = float(total["Scenario Impact"])
    monthly_impact_sum = float(monthly["Monthly Impact"].sum())

    if abs(scenario_impact) < 1e-8:
        flags.append("ZERO_SCENARIO_IMPACT")

    if abs(monthly_impact_sum) < 1e-8 and abs(scenario_impact) > 1e-8:
        flags.append("MONTHLY_IMPACT_ZERO_BUT_ANNUAL_NONZERO")

    if abs(baseline_final - scenario_final) < 1e-8 and abs(scenario_impact) < 1e-8:
        flags.append("BASELINE_EQUALS_SCENARIO")

    if abs(monthly_impact_sum - scenario_impact) > 1e-4:
        flags.append("ANNUAL_MONTHLY_MISMATCH")

    return {
        "Scenario": scenario_name,
        "Selected Year": result["metadata"]["selected_year"],
        "Baseline Final": baseline_final,
        "Scenario Final": scenario_final,
        "Scenario Impact": scenario_impact,
        "Scenario Impact %": float(total["Scenario Impact %"]),
        "Monthly Impact Sum": monthly_impact_sum,
        "Top Gainer 1": top_gain.iloc[0]["Internal Tax Head"] if len(top_gain) > 0 else "",
        "Top Gainer 1 Impact": float(top_gain.iloc[0]["Scenario Impact"]) if len(top_gain) > 0 else 0.0,
        "Top Loser 1": top_loss.iloc[0]["Internal Tax Head"] if len(top_loss) > 0 else "",
        "Top Loser 1 Impact": float(top_loss.iloc[0]["Scenario Impact"]) if len(top_loss) > 0 else 0.0,
        "Flags": " | ".join(flags) if flags else "OK",
    }


def run_sweep(workbook_path: str, selected_year: str = "2025/26"):
    scenarios = get_scenarios(workbook_path)

    rows = []
    failures = []

    for scen in scenarios:
        print(f"Running scenario: {scen}")
        try:
            result = run_one(workbook_path, selected_year, scen)
            rows.append(summarize_result(result, scen))
        except Exception as e:
            failures.append({
                "Scenario": scen,
                "Error": str(e),
                "Traceback": traceback.format_exc(),
            })

    summary_df = pd.DataFrame(rows)
    failures_df = pd.DataFrame(failures)

    if not summary_df.empty:
        summary_df = summary_df.sort_values("Scenario Impact").reset_index(drop=True)

    return summary_df, failures_df


if __name__ == "__main__":
    summary, failures = run_sweep(WORKBOOK, selected_year="2025/26")

    print("\n" + "=" * 100)
    print("SCENARIO SWEEP SUMMARY")
    print("=" * 100)
    print(summary)

    print("\n" + "=" * 100)
    print("FAILED SCENARIOS")
    print("=" * 100)
    if failures.empty:
        print("None")
    else:
        print(failures[["Scenario", "Error"]])

    out_xlsx = Path("scenario_sweep_results.xlsx").resolve()
    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Summary", index=False)
        failures.to_excel(writer, sheet_name="Failures", index=False)

    print(f"\nSaved: {out_xlsx}")